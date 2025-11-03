import json
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.contrib.sites import requests
from django.core import signing
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.core.signing import Signer
from django.db.models import Q, OuterRef, Exists
from django.http import HttpResponseBadRequest, JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET

from .models import UserList, UserRole, FieldMaster, FieldMasterValue, MenuItem, DynamicFormData, LeadTable, ZoneTable, \
    SalesInfoTable, HistorySalesInfo, HistoryLead, BrandTable, CallDisposition, SalesContact, TBLFollowUp, SalesVOC
from .views import render_menu
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from datetime import datetime


@login_required
def sales_user(request):
    # Render Menu
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    # Use email-based matching instead of user FK
    try:
        user_obj = UserList.objects.get(email_id=request.user.email, user_role='Sales')
    except UserList.DoesNotExist:
        return render(request, 'crmapp/sales.html', {
            'menu_html': menu_html,
            'leads': [],
            'zones': [],
            'error': "You are not authorized to view sales leads."
        })

    # Filter leads assigned to this sales user's email
    leads = LeadTable.objects.filter(seller_email_id=user_obj.email_id).order_by('-created_at')

    # Optional: Date filter
    query = request.GET.get('query')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply date filters
    if start_date:
        leads = leads.filter(lead_date__gte=parse_date(start_date))
    if end_date:
        leads = leads.filter(lead_date__lte=parse_date(end_date))

    if query:
        leads = leads.filter(
            Q(customer_name__icontains=query) |
            Q(calling_number__icontains=query) |
            Q(enquiry_type__icontains=query) |
            Q(enquiry_source__icontains=query)
        )

    # Apply pagination
    paginator = Paginator(leads, 10)  # 10 leads per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Zone list for filter dropdowns
    zones = ZoneTable.objects.values_list('zone', flat=True).distinct()

    return render(request, 'crmapp/sales.html', {
        'menu_html': menu_html,
        'leads': page_obj,
        'zones': zones,
        'paginator': paginator,
        'page_obj': page_obj,
    })


def sales_get_data(request):
    encrypted_data = request.GET.get('data')
    if not encrypted_data:
        message = "Missing encrypted data in the URL."
        return render(request, 'crmapp/sales_get_data.html', {'message': message})

    try:
        data = signing.loads(encrypted_data)
        uid = data.get('uid')
        email = data.get('email')
    except signing.BadSignature:
        message = "Invalid or tampered URL."
        return render(request, 'crmapp/sales_get_data.html', {'message': message})

    # Validate presence of uid and email
    if not uid:
        return render(request, 'crmapp/sales_get_data.html', {'message': "UID is missing in the encrypted data."})

    # if not email:
    #     return render(request, 'crmapp/sales_get_data.html', {'message': "Email is missing in the encrypted data."})

    # Query only if both uid and email are present
    lead = LeadTable.objects.filter(id=uid).first()

    if not lead:
        message = "No matching lead found for the given ID and email."
        return render(request, 'crmapp/sales_get_data.html', {'message': message})

    sales_info = SalesInfoTable.objects.filter(lead_table=lead).first()


    return render(request, 'crmapp/sales_get_data.html', {'lead': lead, 'sales': sales_info})

@csrf_exempt
def update_sales_info(request):
    if request.method == 'POST':
        lead_id = request.POST.get('lead_id')
        lead = get_object_or_404(LeadTable, id=lead_id)

        sales_info, created = SalesInfoTable.objects.get_or_create(lead_table=lead)

        sales_info.sale_mt = request.POST.get('sale_mt')
        sales_info.sale_inr = request.POST.get('sale_inr')
        sales_info.sale_team_remarks = request.POST.get('sale_team_remarks')
        sales_info.lead_status = request.POST.get('lead_status')
        sales_info.cc_final_remarks_reformat = request.POST.get('cc_final_remarks_reformat')
        sales_info.lead_category = request.POST.get('lead_category')
        sales_info.product = request.POST.get('product')
        sales_info.product_value = request.POST.get('product_value')
        sales_info.status = request.POST.get('status')

        # ✅ Only assign user if authenticated
        # if created and request.user.is_authenticated:
        #     sales_info.created_by = request.user
        # if request.user.is_authenticated:
        #     sales_info.updated_by = request.user

        sales_info.save()

        # Save to history table (optional: add created_by if needed)
        HistorySalesInfo.objects.create(
            lead_table=sales_info,
            sale_mt=sales_info.sale_mt,
            sale_inr=sales_info.sale_inr,
            sale_team_remarks=sales_info.sale_team_remarks,
            lead_status=sales_info.lead_status,
            cc_final_remarks_reformat=sales_info.cc_final_remarks_reformat,
            lead_category=sales_info.lead_category,
            product=sales_info.product,
            product_value=sales_info.product_value,
            status=sales_info.status,
        )

        # ✅ Encrypt and redirect
        encrypted_data = signing.dumps({
            'uid': lead.id,
            'email': lead.seller_email_id
        })

        return redirect(f'/sales_get_data/?data={encrypted_data}')

    return redirect('dashboard')

from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
@login_required
def lead_detail(request, lead_id):
    state = request.GET.get('state', '')
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    # enquiry_sources = [
    #     "ChatBot", "Exhibition", "GetDistributor", "Inbound", "IndiaMart", "Mail", "Meta",
    #     "Plantix", "TradeIndia", "WebScraping", "WebSite", "Other"
    # ]

    lead = get_object_or_404(LeadTable, id=lead_id)
    payload = {'uid': lead.id, 'email': lead.seller_email_id}
    encrypted_data = signing.dumps(payload)
    secure_url = request.build_absolute_uri(f"/sales_get_data/?data={encrypted_data}")

    # 🔹 Paginate Lead History
    lead_history_qs = HistoryLead.objects.filter(lead_table_id=lead_id).order_by('-created_at')
    page_lead = request.GET.get('page_lead')

    paginator_lead = Paginator(lead_history_qs, 5)
    try:
        lead_history = paginator_lead.page(page_lead)
    except PageNotAnInteger:
        lead_history = paginator_lead.page(1)
    except EmptyPage:
        lead_history = paginator_lead.page(paginator_lead.num_pages)

    # 🔹 Sales Info (not paginated because it's one-to-one)
    sales_info = SalesInfoTable.objects.filter(lead_table=lead).first()

    # 🔹 Paginate Sales History
    sales_history_qs = HistorySalesInfo.objects.filter(lead_table=sales_info) if sales_info else []
    page_sales = request.GET.get('page_sales')
    paginator_sales = Paginator(sales_history_qs, 10)
    try:
        sales_history = paginator_sales.page(page_sales)
    except PageNotAnInteger:
        sales_history = paginator_sales.page(1)
    except EmptyPage:
        sales_history = paginator_sales.page(paginator_sales.num_pages)

        # 🔹 Brand/Product/SubProduct Dropdown Data
    brand_data_qs = BrandTable.objects.all().values('brand', 'product_types', 'sub_product_types')
    brand_data = list(brand_data_qs)  # Convert queryset to list of dicts
    brands = sorted(set(item['brand'] for item in brand_data if item['brand']))  # Unique brands

    # Get call disposition data
    call_dispositions_qs = CallDisposition.objects.all().values('type', 'sub_type', 'sub_sub_type')
    call_dispositions = list(call_dispositions_qs)  # to use in JS

    lead = get_object_or_404(LeadTable, id=lead_id)
    state = lead.state

    sales_contacts = SalesContact.objects.filter(state__iexact=state).values(
        'city_custom', 'l1_name', 'l1_mail', 'l1_mobile'
    )

    # Convert to list if needed
    sales_contacts = list(sales_contacts)

    # Fetch follow-up history for this lead
    followup_qs = TBLFollowUp.objects.filter(lead_table=lead).all()
    page_followup = request.GET.get('page_followup')

    paginator_followup = Paginator(followup_qs, 5)  # Show 5 per page
    try:
        followup_history = paginator_followup.page(page_followup)
    except PageNotAnInteger:
        followup_history = paginator_followup.page(1)
    except EmptyPage:
        followup_history = paginator_followup.page(paginator_followup.num_pages)

    now = timezone.now()

    enquiry_source_mapping = {
        "ChatBot": ["WhatsApp"],
        "Exhibition": ["Exhibition", "MPSO Exhibition", "Poultry Expo", "Rachana Visitors", "Sthapatya Exhibition"],
        "GetDistributor": ["BuyLeads", "Direct Leads"],
        "Inbound": ["Banner", "BTL", "Coupon", "Facebook", "birlanu.com", "IndiaMart SMS", "Info Mail", "Instagram",
                    "SMS", "Word of Mouth", "YouTube"],
        "IndiaMart": ["BuyLeads", "Direct Leads", "IndiaMart Mail"],
        "Mail": ["info@hil.in","Customer Care"],
        "Plantix": ["Direct Leads"],
        "Meta": ["Facebook", "Instagram"],
        "Social": ["ORM"],
        "TradeIndia": ["Direct Leads"],
        "WebScraping": ["Walling Project"],
        "Website": ["Contact", "Pop-up"],
        "Other": []
    }

    # followup_count = lead.follow_lead_table.count()
    # if lead.sub_calling_status == "Invalid":
    #     # Highest priority: invalid case
    #     lead.lead_closer_status = "invalid_close"
    # else:
    #     # Fallback: use follow-up logic
    #
    #
    #     if followup_count == 0:
    #         lead.lead_closer_status = "pending"
    #     elif 1 <= followup_count <= 4:
    #         lead.lead_closer_status = f"attempted_{followup_count}"
    #     else:
    #         lead.lead_closer_status = "no_response"

    return render(request, 'crmapp/lead_detail.html', {
        'lead': lead,
        'menu_html': menu_html,
        "enquiry_sources": list(enquiry_source_mapping.keys()),
        "enquiry_source_mapping": enquiry_source_mapping,

        'secure_url': secure_url,
        'lead_history': lead_history,
        'sales_info': sales_info,
        'sales_history': sales_history,
        'brands': brands,  # list of brand strings
        'brand_data': brand_data,  # queryset or list of dicts
        'call_dispositions': call_dispositions,
        'sales_contacts': sales_contacts,
        'followup_history': followup_history,
        'current_time': now,
        # 'followup_count': followup_count,
    })



def get_zone_data(request):
    pincode = request.GET.get('pincode')

    if not pincode:
        return JsonResponse({'status': 'error', 'message': 'Pincode is required'}, status=400)

    try:
        zone_info = ZoneTable.objects.filter(pincode=pincode).first()

        print(f'SELECT * FROM zone_table WHERE pincode="{pincode}" LIMIT 1;')

        if zone_info:
            state = zone_info.state_ut or ''

            # Fetch sales contacts for the detected state
            sales_contacts = list(
                SalesContact.objects.filter(state__iexact=state).values(
                    'city_custom', 'l1_name', 'l1_mail', 'l1_mobile'
                )
            )

            return JsonResponse({
                'status': 'success',
                'district': zone_info.district or '',
                'state': state,
                'zone': zone_info.zone or '',
                'sales_contacts': sales_contacts,
            })

        else:
            return JsonResponse({'status': 'not_found', 'message': 'No data found for this pincode'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)




@login_required
def save_follow_up(request, lead_id):
    if request.method == 'POST':
        lead = get_object_or_404(LeadTable, id=lead_id)
        status = request.POST.get('status')
        sub_status = request.POST.get('connection_status')
        sales_person_voc = request.POST.get('sales_person_voc')
        remark = request.POST.get('remark')
        followup_time = request.POST.get('FollowUp_time')
        followup_status= "0" if followup_time else "1"
        follow_up_from = request.POST.get('follow_up')  # should be 'customer' or 'seller'

        TBLFollowUp.objects.create(
            lead_table=lead,
            status=status,
            sub_status=sub_status,
            sales_voc=sales_person_voc,
            remark=remark,
            followup_time=followup_time if followup_time else None,
            followup_status=followup_status,
            follow_up=follow_up_from,
            created_by=request.user,
            updated_by=request.user,
        )
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

import requests
def make_call_api(request):
    user = request.user
    try:
        user_list_entry = UserList.objects.get(user=user)
        username = user_list_entry.username
    except UserList.DoesNotExist:
        return JsonResponse({'error': 'UserList entry not found for current user.'}, status=404)

    number = request.GET.get('number')


    if not number:
        return JsonResponse({'status': 'error', 'message': 'Number is required'})

    # Prepare API parameters
    api_url = "http://192.168.11.6/C2Capi/api_birla.php"
    params = {
        "customer_number": number,
        "agent_user": username,
        "token": "VHJoc2xkZ2dkXjc1MzYzNVVVR2hzZ3M2",
    }

    try:
        response = requests.get(api_url, params=params)
        response_data = response.json()


        return JsonResponse({'status': 'success', 'api_response': response_data})

    except requests.exceptions.RequestException as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    except ValueError:
        # If response is not JSON
        return JsonResponse({'status': 'error', 'message': 'Invalid response from API'})

import openpyxl
@login_required
def leads_export(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if from_date_str and to_date_str:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)

        if not from_date or not to_date:
            return render(request, 'crmapp/leads_export.html', {
                "menu_html": menu_html,
                "error": "Invalid date format."
            })

        from_datetime = datetime.combine(from_date, datetime.min.time())
        to_datetime = datetime.combine(to_date, datetime.max.time())

        leads = HistoryLead.objects.filter(created_at__range=(from_datetime, to_datetime))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        ws.append([
            'ID',
            'Customer Name', 'Customer Type', 'Calling Number', 'Enquiry Type', 'Enquiry Source', 'Sub Enquiry Source',
            'Lead Date', 'Call Date', 'Call Type', 'Calling Status', 'Interested Status',
            'Sub Calling Status', 'Sub Sub Calling Status', 'Select Business', 'Buyer Type',
            'Lead Status', 'Construction Level', 'Name', 'Alternative Number', 'Email ID',
            'Address', 'Landmark', 'Brand', 'Product', 'Sub Product',
            'State', 'District', 'Zone', 'Pincode', 'Agent Name',
            'Order Qty', 'Order Description', 'Order Value', 'Customer Type Select',
            'Registration Status', 'Remark', 'Secure URL', 'Seller Email ID', 'Seller Phone No','Lead Closer Status',
            'Lead Closer Status New',
            'Created By', 'Updated By', 'Created At', 'Updated At'
        ])

        for idx, lead in enumerate(leads, start=1):
            ws.append([
                idx,
                lead.customer_name,
                lead.customer_type,
                lead.calling_number,
                lead.enquiry_type,
                lead.enquiry_source,
                lead.sub_enquiry_source,
                lead.lead_date.strftime('%Y-%m-%d') if lead.lead_date else '',
                lead.call_date.strftime('%Y-%m-%d') if lead.call_date else '',
                lead.call_type,
                lead.calling_status,
                lead.interested_status,
                lead.sub_calling_status,
                lead.sub_sub_calling_status,
                lead.select_bus,
                lead.buyer_type,
                lead.lead_status,
                lead.construction_level,
                lead.name,
                lead.alternative_number,
                lead.email_id,
                lead.address,
                lead.landmark,
                lead.brand,
                lead.product,
                lead.sub_product,
                lead.state,
                lead.district,
                lead.zone,
                lead.pin_code,
                str(lead.created_by).split('@')[0] if lead.created_by else '',
                lead.order_qty,
                lead.order_description,
                float(lead.order_value) if lead.order_value else '',
                lead.customer_type_select,
                lead.registration_status,
                lead.remark,
                lead.secure_url,
                lead.seller_email_id,
                lead.seller_phone_no,
                lead.lead_closer_status,
                lead.lead_closer_status_new,
                str(lead.created_by) if lead.created_by else '',
                str(lead.updated_by) if lead.updated_by else '',
                lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '',
                lead.updated_at.strftime('%Y-%m-%d %H:%M:%S') if lead.updated_at else ''
            ])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return FileResponse(
            file_stream,
            as_attachment=True,
            filename='history_leads.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # No export, show template
    return render(request, 'crmapp/leads_export.html', {
        "menu_html": menu_html
    })

@login_required
def sales_export(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if from_date_str and to_date_str:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)

        if not from_date or not to_date:
            return render(request, 'crmapp/sales_export.html', {
                "menu_html": menu_html,
                "error": "Invalid date format."
            })
        from_datetime = datetime.combine(from_date, datetime.min.time())
        to_datetime = datetime.combine(to_date, datetime.max.time())

        leads = HistorySalesInfo.objects.filter(
            created_at__range=(from_datetime, to_datetime)
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Leads"

        # Header row
        ws.append([
            'ID',
            'Sale MT', 'Sale INR', 'Sale Team Remarks', 'Lead Status',
            'CC Final Remarks Reformat', 'Lead Category', 'Product',
            'Product Value', 'Status', 'Created By', 'Updated By',
            'Created At', 'Updated At'
        ])

        # Data rows
        for idx, lead in enumerate(leads, start=1):
            ws.append([
                idx,
                lead.sale_mt,
                lead.sale_inr,
                lead.sale_team_remarks,
                lead.lead_status,
                lead.cc_final_remarks_reformat,
                lead.lead_category,
                lead.product,
                lead.product_value,
                lead.status,
                str(lead.created_by) if lead.created_by else '',
                str(lead.updated_by) if lead.updated_by else '',
                lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '',
                lead.updated_at.strftime('%Y-%m-%d %H:%M:%S') if lead.updated_at else ''
            ])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return FileResponse(
            file_stream,
            as_attachment=True,
            filename='sales_history_leads.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # Show template
    return render(request, 'crmapp/sales_export.html', {
        "menu_html": menu_html
    })

@login_required
def follow_up(request):
    # Build the menu
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        menu_tree.setdefault(item.parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    # Get dates
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if from_date_str and to_date_str:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)

        if not from_date or not to_date:
            return render(request, 'crmapp/follow_up.html', {
                "menu_html": menu_html,
                "error": "Invalid date format."
            })

        from_datetime = datetime.combine(from_date, datetime.min.time())
        to_datetime = datetime.combine(to_date, datetime.max.time())

        followups = TBLFollowUp.objects.filter(
            created_at__range=(from_datetime, to_datetime)
        )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Follow Ups"

        # Header
        ws.append([
            'ID',
            'Lead ID', 'Status', 'Sub Status', 'Remark', 'Follow Up',
            'Created By', 'Updated By', 'Created At'
        ])

        # Data rows
        for idx, entry in enumerate(followups, start=1):
            ws.append([
                idx,
                entry.lead_table.id if entry.lead_table else '',
                entry.status,
                entry.sub_status,
                entry.remark,
                entry.follow_up,
                entry.created_by.username if entry.created_by else '',
                entry.updated_by.username if entry.updated_by else '',
                entry.created_at.strftime('%Y-%m-%d %H:%M:%S') if entry.created_at else '',
            ])

        # Save to in-memory file
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return FileResponse(
            file_stream,
            as_attachment=True,
            filename='follow_up_export.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # Initial load without export
    return render(request, 'crmapp/follow_up.html', {
        "menu_html": menu_html
    })




@login_required
def main_leads_export(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if from_date_str and to_date_str:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)

        if not from_date or not to_date:
            return render(request, 'crmapp/leads_export.html', {
                "menu_html": menu_html,
                "error": "Invalid date format."
            })

        from_datetime = datetime.combine(from_date, datetime.min.time())
        to_datetime = datetime.combine(to_date, datetime.max.time())

        leads = LeadTable.objects.filter(created_at__range=(from_datetime, to_datetime))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        ws.append([
            'ID',
            'Unique Id',
            'Customer Name', 'Customer Type', 'Calling Number', 'Enquiry Type', 'Enquiry Source', 'Sub Enquiry Source',
            'Lead Date', 'Call Date', 'Call Type', 'Calling Status', 'Interested Status',
            'Sub Calling Status', 'Sub Sub Calling Status', 'Select Business', 'Buyer Type',
            'Lead Status', 'Construction Level', 'Name', 'Alternative Number', 'Email ID',
            'Address', 'Landmark', 'Brand', 'Product', 'Sub Product',
            'State', 'District', 'Zone', 'Pincode', 'Agent Name',
            'Order Qty', 'Order Description', 'Order Value', 'Customer Type Select',
            'Registration Status', 'Remark', 'Secure URL', 'Sales Officer Email Id', 'Sales Officer Contact Number','Lead Closer Status',
            'Lead Closer Status New',
            'Lead Close Date',
            'Final Lead Close Date',
            'Lead Upload Type',
            'Created By', 'Updated By', 'Created At', 'Updated At',

            # --- Sales Info ---
            'Sale MT', 'Sale INR', 'Sale Team Remarks',
            'Sale Lead Status', 'CC Final Remarks',
            'Sale Lead Category', 'Sale Product',
            'Sale Product Value', 'Sale Status'
        ])

        for idx, lead in enumerate(leads, start=1):
            sales = SalesInfoTable.objects.filter(lead_table=lead).first()
            ws.append([
                idx,
                f"BN{lead.pk}",
                lead.customer_name,
                lead.customer_type,
                lead.calling_number,
                lead.enquiry_type,
                lead.enquiry_source,
                lead.sub_enquiry_source,
                lead.lead_date.strftime('%Y-%m-%d') if lead.lead_date else '',
                lead.call_date.strftime('%Y-%m-%d') if lead.call_date else '',
                lead.call_type,
                lead.calling_status,
                lead.interested_status,
                lead.sub_calling_status,
                lead.sub_sub_calling_status,
                lead.select_bus,
                lead.buyer_type,
                lead.lead_status,
                lead.construction_level,
                lead.name,
                lead.alternative_number,
                lead.email_id,
                lead.address,
                lead.landmark,
                lead.brand,
                lead.product,
                lead.sub_product,
                lead.state,
                lead.district,
                lead.zone,
                lead.pin_code,
                str(lead.created_by).split('@')[0] if lead.created_by else '',
                lead.order_qty,
                lead.order_description,
                float(lead.order_value) if lead.order_value else '',
                lead.customer_type_select,
                lead.registration_status,
                lead.remark,
                lead.secure_url,
                lead.seller_email_id,
                lead.seller_phone_no,
                lead.lead_closer_status,
                lead.lead_closer_status_new,
                lead.lead_close_date,
                lead.final_lead_close_date,
                lead.lead_upload_type,
                str(lead.created_by) if lead.created_by else '',
                str(lead.updated_by) if lead.updated_by else '',
                lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '',
                lead.updated_at.strftime('%Y-%m-%d %H:%M:%S') if lead.updated_at else '',

                sales.sale_mt if sales else '',
                sales.sale_inr if sales else '',
                sales.sale_team_remarks if sales else '',
                sales.lead_status if sales else '',
                sales.cc_final_remarks_reformat if sales else '',
                sales.lead_category if sales else '',
                sales.product if sales else '',
                sales.product_value if sales else '',
                sales.status if sales else '',
            ])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return FileResponse(
            file_stream,
            as_attachment=True,
            filename='leads_data.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # No export, show template
    return render(request, 'crmapp/leads_export.html', {
        "menu_html": menu_html
    })



@login_required
def reallocate(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    users = User.objects.all()
    leads = LeadTable.objects.all()

    if request.method == 'POST':
        from_user_id = request.POST.get('from_user')
        to_user_id = request.POST.get('to_user')
        lead_id = request.POST.get('lead_id')

        try:
            from_user = User.objects.get(id=from_user_id)
            to_user = User.objects.get(id=to_user_id)

            if lead_id:
                # Transfer specific lead
                lead = get_object_or_404(LeadTable, id=lead_id, created_by=from_user)
                lead.created_by = to_user
                lead.lead_action = f"{from_user.get_full_name() or from_user.username} transferred the lead to {to_user.get_full_name() or to_user.username}"
                lead.save()

                # Save history (only lead_action)
                HistoryLead.objects.create(
                    lead_table=lead,
                    lead_action=lead.lead_action,
                    created_by=from_user,
                    updated_by=to_user
                )

                messages.success(
                    request,
                    f" Lead {lead.id} successfully reallocated from {from_user.email} to {to_user.email}."
                )
            else:
                # Transfer all leads from from_user to to_user
                leads_to_update = LeadTable.objects.filter(created_by=from_user)
                for lead in leads_to_update:
                    lead.created_by = to_user
                    lead.lead_action = f"{from_user.get_full_name() or from_user.username} transferred the lead to {to_user.get_full_name() or to_user.username}"
                    lead.save()

                    # Save history (only lead_action)
                    HistoryLead.objects.create(
                        lead_table=lead,
                        lead_action=lead.lead_action,
                        created_by=from_user,
                        updated_by=to_user
                    )

                messages.success(
                    request,
                    f" {leads_to_update.count()} leads successfully transferred from {from_user.email} to {to_user.email}."
                )

        except User.DoesNotExist:
            messages.error(request, " One or both users not found.")
        except LeadTable.DoesNotExist:
            messages.error(request, " Lead not found or does not belong to selected From User.")
        except Exception as e:
            messages.error(request, f" Error: {str(e)}")

        return redirect('reallocate')

    return render(request, 'crmapp/reallocate.html', {
        'users': users,
        'leads': leads,
        'menu_html': menu_html,
    })



@require_GET
def get_voc_options_api(request):
    filtered_qs = SalesVOC.objects.filter(followup_for="sales").values(
        'information_status', 'connection_status', 'sales_person_voc'
    )

    # Return the list of records as-is
    return JsonResponse({
        'status': 'success',
        'records': list(filtered_qs)
    })


@require_GET
def get_customer_voc_options_api(request):
    filtered_qs = SalesVOC.objects.filter(followup_for="customer").values(
        'information_status', 'connection_status', 'sales_person_voc'
    )

    return JsonResponse({
        'status': 'success',
        'records': list(filtered_qs)
    })


@csrf_exempt
def copy_lead(request, lead_id):
    if request.method == "POST":
        try:
            lead = LeadTable.objects.get(id=lead_id)

            # Create a new instance with copied data
            new_lead = LeadTable.objects.create(
                customer_name=lead.customer_name,
                calling_number=lead.calling_number,
                enquiry_type=lead.enquiry_type,
                enquiry_source=lead.enquiry_source,
                lead_date=lead.lead_date,
                created_by=request.user if request.user.is_authenticated else None
            )

            return JsonResponse({'status': 'success', 'new_id': new_lead.id})
        except LeadTable.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Lead not found'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



def bulk_upload(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    return render(request, 'crmapp/bulk_upload.html', {
        'menu_html': menu_html,
    })

import pandas as pd
@csrf_exempt
def upload_leads_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']

        try:
            df = pd.read_excel(excel_file)

            success_count = 0
            failed_entries = []

            for _, row in df.iterrows():
                try:
                    lead_id = str(row.get('lead_id', '')).strip()  # Expected to be User ID
                    customer_name = str(row.get('customer_name', '')).strip()
                    customer_type = str(row.get('Campaign_type', '')).strip()
                    calling_number = str(row.get('calling_number', '')).strip()
                    enquiry_type = str(row.get('enquiry_type', '')).strip()
                    enquiry_source = str(row.get('enquiry_source', '')).strip()
                    lead_date_str = row.get('lead_date', None)

                    lead_date = (
                        pd.to_datetime(lead_date_str).date()
                        if not pd.isna(lead_date_str)
                        else None
                    )

                    # Try to get User from lead_id
                    try:
                        user = User.objects.get(id=int(lead_id))
                    except (User.DoesNotExist, ValueError):
                        failed_entries.append({
                            'calling_number': calling_number,
                            'reason': f'User with ID {lead_id} does not exist'
                        })
                        continue

                    # Check for duplicates and block if necessary
                    matching_leads = LeadTable.objects.filter(calling_number=calling_number)

                    block = False
                    for lead in matching_leads:
                        sales_qs = SalesInfoTable.objects.filter(lead_table=lead)
                        if sales_qs.exists():
                            if not sales_qs.filter(status__iexact='closed').exists():
                                block = True
                                break
                        else:
                            block = True
                            break

                    if block:
                        failed_entries.append({
                            'calling_number': calling_number,
                            'reason': 'Lead already exists and is not closed'
                        })
                        continue

                    # Create Lead
                    LeadTable.objects.create(
                        customer_name=customer_name,
                        customer_type=customer_type,
                        calling_number=calling_number,
                        enquiry_type=enquiry_type,
                        enquiry_source=enquiry_source,
                        lead_date=lead_date,
                        lead_upload_type="Bulk",
                        created_by=user
                    )
                    success_count += 1

                except Exception as e:
                    failed_entries.append({
                        'calling_number': row.get('calling_number'),
                        'reason': str(e)
                    })

            return JsonResponse({
                'status': 'completed',
                'success_count': success_count,
                'failed_entries': failed_entries
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'message': 'Invalid request'}, status=400)

def download_excel_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads Template"

    # Add headers
    headers = ['lead_id', 'customer_name','Campaign_type', 'calling_number', 'enquiry_type', 'enquiry_source', 'lead_date']
    ws.append(headers)

    ws.append(['15', 'Sachin kr','Self', '9876543210', 'Product Inquiry', 'Website', '2025-07-14'])

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="lead_template.xlsx"'
    return response


@login_required
def call_back_lead(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    # Base leads for current user
    leads = LeadTable.objects.filter(created_by=request.user).order_by('-created_at')

    # Get filters
    query = request.GET.get('query')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    calling_status = request.GET.get('calling_status')
    sub_calling_status = request.GET.get('sub_calling_status')

    # Apply filters
    if calling_status:
        leads = leads.filter(calling_status__iexact=calling_status)

    if sub_calling_status:
        leads = leads.filter(sub_calling_status__iexact=sub_calling_status)
    else:
        leads = leads.filter(sub_calling_status__iexact='Call Back')  # default

    if start_date:
        leads = leads.filter(lead_date__gte=parse_date(start_date))

    if end_date:
        leads = leads.filter(lead_date__lte=parse_date(end_date))

    if query:
        leads = leads.filter(
            Q(customer_name__icontains=query) |
            Q(calling_number__icontains=query) |
            Q(enquiry_type__icontains=query) |
            Q(enquiry_source__icontains=query) |
            Q(sub_calling_status__icontains=query)
        )

    # Pagination
    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Zones list (optional)
    zones = ZoneTable.objects.values_list('zone', flat=True).distinct()

    return render(request, 'crmapp/call_back_lead.html', {
        'menu_html': menu_html,
        'leads': page_obj,
        'zones': zones,
        'paginator': paginator,
        'page_obj': page_obj,
    })



@login_required
def follow_up_data(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    followup_with_time = TBLFollowUp.objects.filter(
        lead_table=OuterRef('pk'),
        followup_time__isnull=False
    )

    leads = LeadTable.objects.filter(
        created_by=request.user
    ).annotate(
        has_followup_time=Exists(followup_with_time)
    ).filter(
        has_followup_time=True
    ).order_by('-created_at')

    # Filters
    query = request.GET.get('query')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    calling_status = request.GET.get('calling_status')
    sub_calling_status = request.GET.get('sub_calling_status')

    if calling_status:
        leads = leads.filter(calling_status__iexact=calling_status)

    if sub_calling_status:
        leads = leads.filter(sub_calling_status__iexact=sub_calling_status)

    if start_date:
        leads = leads.filter(lead_date__gte=parse_date(start_date))

    if end_date:
        leads = leads.filter(lead_date__lte=parse_date(end_date))

    if query:
        leads = leads.filter(
            Q(customer_name__icontains=query) |
            Q(calling_number__icontains=query) |
            Q(enquiry_type__icontains=query) |
            Q(enquiry_source__icontains=query) |
            Q(sub_calling_status__icontains=query)
        )

    # Pagination
    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    zones = ZoneTable.objects.values_list('zone', flat=True).distinct()

    return render(request, 'crmapp/follow_up_data.html', {
        'menu_html': menu_html,
        'leads': page_obj,
        'zones': zones,
        'paginator': paginator,
        'page_obj': page_obj,
    })


@login_required
def get_leads_by_user(request, user_id):
    leads = LeadTable.objects.filter(created_by_id=user_id).values('id', 'customer_name','enquiry_source')
    return JsonResponse(list(leads), safe=False)


from django.template.loader import render_to_string
def send_lead_email(request, lead_id):
    try:
        # Fetch the lead
        lead = LeadTable.objects.get(id=lead_id)

        # Use the existing secure_url from the DB
        if not lead.secure_url:
            return HttpResponse("Secure URL not available for this lead.", status=400)

        secure_url = request.build_absolute_uri(lead.secure_url)

        # Prepare context for the email template
        context = {
            "lead": lead,
            "secure_url": secure_url,
            "formatted_date": lead.callback_time.strftime("%d %B %Y, %I:%M %p") if lead.callback_time else "",
        }

        # Render email content
        html_content = render_to_string("crmapp/lead_detail_email.html", context)
        text_content = "Lead Details - Please open in an HTML-supported email viewer."

        # Email details
        subject = "Validated Lead Information"
        from_email = settings.EMAIL_HOST_USER
        to = [lead.seller_email_id]

        # Ensure cc is a list or empty list
        cc = [lead.seller_email_id_L2] if lead.seller_email_id_L2 else []

        # Create and send the email
        msg = EmailMultiAlternatives(subject, text_content, from_email, to, cc=cc)
        msg.attach_alternative(html_content, "text/html")
        msg.send()

        return HttpResponse("Lead email sent successfully.")

    except LeadTable.DoesNotExist:
        return HttpResponse("Lead not found.", status=404)



@login_required
def not_connected_lead(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    # Base leads for current user
    leads = LeadTable.objects.filter(created_by=request.user).order_by('-created_at')

    # Get filters
    query = request.GET.get('query')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    calling_status = request.GET.get('calling_status')
    sub_calling_status = request.GET.get('sub_calling_status')

    # Apply filters
    if calling_status:
        leads = leads.filter(calling_status__iexact=calling_status)
    else:
        leads = leads.filter(calling_status__iexact='Not Connect')

    if sub_calling_status:
        leads = leads.filter(sub_calling_status__iexact=sub_calling_status)

    if start_date:
        leads = leads.filter(lead_date__gte=parse_date(start_date))

    if end_date:
        leads = leads.filter(lead_date__lte=parse_date(end_date))

    if query:
        leads = leads.filter(
            Q(customer_name__icontains=query) |
            Q(calling_number__icontains=query) |
            Q(enquiry_type__icontains=query) |
            Q(enquiry_source__icontains=query) |
            Q(sub_calling_status__icontains=query)
        )

    # Pagination
    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Zones list (optional)
    zones = ZoneTable.objects.values_list('zone', flat=True).distinct()

    return render(request, 'crmapp/not_connect_lead.html', {
        'menu_html': menu_html,
        'leads': page_obj,
        'zones': zones,
        'paginator': paginator,
        'page_obj': page_obj,
    })


@login_required
def new_lead(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    # Base leads for current user
    leads = LeadTable.objects.filter(created_by=request.user).order_by('-created_at')

    # Get filters
    query = request.GET.get('query')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    calling_status = request.GET.get('calling_status')
    sub_calling_status = request.GET.get('sub_calling_status')

    # Apply filters
    if calling_status:
        leads = leads.filter(calling_status__iexact=calling_status)

    if sub_calling_status:
        leads = leads.filter(sub_calling_status__iexact=sub_calling_status)

    if start_date:
        leads = leads.filter(lead_date__gte=parse_date(start_date))

    if end_date:
        leads = leads.filter(lead_date__lte=parse_date(end_date))

    if query:
        leads = leads.filter(
            Q(customer_name__icontains=query) |
            Q(calling_number__icontains=query) |
            Q(enquiry_type__icontains=query) |
            Q(enquiry_source__icontains=query) |
            Q(sub_calling_status__icontains=query)
        )
    leads = leads.filter(Q(secure_url__isnull=True) | Q(secure_url__exact='')).order_by('-created_at')

    # Pagination
    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Zones list (optional)
    zones = ZoneTable.objects.values_list('zone', flat=True).distinct()

    return render(request, 'crmapp/new_lead.html', {
        'menu_html': menu_html,
        'leads': page_obj,
        'zones': zones,
        'paginator': paginator,
        'page_obj': page_obj,
    })


########## whatsapp view api  ##########

@csrf_exempt
def send_whatsapp_message(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            phone = data.get("phone")
            template_id = data.get("templateId", "purchase_query")


            payload = {
                "message": {
                    "channel": "WABA",
                    "content": {
                        "preview_url": False,
                        "type": "TEMPLATE",
                        "template": {
                            "templateId": template_id,
                            "parameterValues": {}  # Use if template needs dynamic values
                        },
                        "shorten_url": True
                    },
                    "recipient": {
                        "to": phone,
                        "recipient_type": "individual",
                        "reference": {
                            "cust_ref": "Some Customer Ref",
                            "messageTag1": "Message Tag Val1",
                            "conversationId": "Some Optional Conversation ID"
                        }
                    },
                    "sender": {
                        "from": "918977754652"
                    },
                    "preferences": {
                        "webHookDNId": "1001"
                    }
                },
                "metaData": {
                    "version": "v1.0.9"
                }
            }

            headers = {
                "Authentication": "Bearer KqHuXlEn8rg1wOv5slUhuA==",
                "Content-Type": "application/json"
            }

            response = requests.post(
                "https://rcmapi.instaalerts.zone/services/rcm/sendMessage",
                headers=headers,
                json=payload
            )

            return JsonResponse(response.json(), status=response.status_code)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)



@login_required
def call_date(request):
    menu_items = MenuItem.objects.filter(is_active=True).order_by('order')
    menu_tree = {}
    for item in menu_items:
        parent_id = item.parent_id
        menu_tree.setdefault(parent_id, []).append(item)
    menu_html = render_menu(None, menu_tree)

    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if from_date_str and to_date_str:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)

        if not from_date or not to_date:
            return render(request, 'crmapp/leads_export.html', {
                "menu_html": menu_html,
                "error": "Invalid date format."
            })

        # from_datetime = datetime.combine(from_date, datetime.min.time())
        # to_datetime = datetime.combine(to_date, datetime.max.time())

        # leads = LeadTable.objects.filter(created_at__range=(from_datetime, to_datetime))
        leads = LeadTable.objects.filter(call_date__range=(from_date, to_date))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        ws.append([
            'ID',
            'Unique Id',
            'Customer Name', 'Customer Type', 'Calling Number', 'Enquiry Type', 'Enquiry Source', 'Sub Enquiry Source',
            'Lead Date', 'Call Date', 'Call Type', 'Calling Status', 'Interested Status',
            'Sub Calling Status', 'Sub Sub Calling Status', 'Select Business', 'Buyer Type',
            'Lead Status', 'Construction Level', 'Name', 'Alternative Number', 'Email ID',
            'Address', 'Landmark', 'Brand', 'Product', 'Sub Product',
            'State', 'District', 'Zone', 'Pincode', 'Agent Name',
            'Order Qty', 'Order Description', 'Order Value', 'Customer Type Select',
            'Registration Status', 'Remark', 'Secure URL', 'Seller Email ID', 'Seller Phone No','Lead Closer Status',
            'Lead Closer Status New',
            'Lead Close Date',
            'Final Lead Close Date',
            'Created By', 'Updated By', 'Created At', 'Updated At',

            # --- Sales Info ---
            'Sale MT', 'Sale INR', 'Sale Team Remarks',
            'Sale Lead Status', 'CC Final Remarks',
            'Sale Lead Category', 'Sale Product',
            'Sale Product Value', 'Sale Status'
        ])

        for idx, lead in enumerate(leads, start=1):
            sales = SalesInfoTable.objects.filter(lead_table=lead).first()
            ws.append([
                idx,
                f"BN{lead.pk}",
                lead.customer_name,
                lead.customer_type,
                lead.calling_number,
                lead.enquiry_type,
                lead.enquiry_source,
                lead.sub_enquiry_source,
                lead.lead_date.strftime('%Y-%m-%d') if lead.lead_date else '',
                lead.call_date.strftime('%Y-%m-%d') if lead.call_date else '',
                lead.call_type,
                lead.calling_status,
                lead.interested_status,
                lead.sub_calling_status,
                lead.sub_sub_calling_status,
                lead.select_bus,
                lead.buyer_type,
                lead.lead_status,
                lead.construction_level,
                lead.name,
                lead.alternative_number,
                lead.email_id,
                lead.address,
                lead.landmark,
                lead.brand,
                lead.product,
                lead.sub_product,
                lead.state,
                lead.district,
                lead.zone,
                lead.pin_code,
                str(lead.created_by).split('@')[0] if lead.created_by else '',
                lead.order_qty,
                lead.order_description,
                float(lead.order_value) if lead.order_value else '',
                lead.customer_type_select,
                lead.registration_status,
                lead.remark,
                lead.secure_url,
                lead.seller_email_id,
                lead.seller_phone_no,
                lead.lead_closer_status,
                lead.lead_closer_status_new,
                lead.lead_close_date,
                lead.final_lead_close_date,
                str(lead.created_by) if lead.created_by else '',
                str(lead.updated_by) if lead.updated_by else '',
                lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '',
                lead.updated_at.strftime('%Y-%m-%d %H:%M:%S') if lead.updated_at else '',

                sales.sale_mt if sales else '',
                sales.sale_inr if sales else '',
                sales.sale_team_remarks if sales else '',
                sales.lead_status if sales else '',
                sales.cc_final_remarks_reformat if sales else '',
                sales.lead_category if sales else '',
                sales.product if sales else '',
                sales.product_value if sales else '',
                sales.status if sales else '',
            ])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return FileResponse(
            file_stream,
            as_attachment=True,
            filename='call_data.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # No export, show template
    return render(request, 'crmapp/call_date.html', {
        "menu_html": menu_html
    })


@csrf_exempt
def upload_leads_update(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']

        try:
            df = pd.read_excel(excel_file)
            success_count = 0
            update_count = 0
            failed_entries = []

            for _, row in df.iterrows():
                try:
                    unique_id = str(row.get('Unique Id', '')).strip()
                    calling_number = str(row.get('Calling Number', '')).strip()
                    customer_name = str(row.get('Customer Name', '')).strip()

                    if not calling_number:
                        failed_entries.append({
                            'Customer Name': customer_name or 'Unknown',
                            'reason': 'Calling Number is missing'
                        })
                        continue

                    # Try to find existing lead by Unique ID
                    lead = None
                    if unique_id:
                        try:
                            numeric_id = int(''.join(filter(str.isdigit, unique_id)))
                            lead = LeadTable.objects.filter(id=numeric_id).first()
                        except ValueError:
                            lead = None

                    # --- Update existing lead ---
                    if lead:
                        # Update all fields if present
                        lead.calling_number = calling_number
                        lead.customer_name = customer_name
                        lead.customer_type = row.get('Customer Type') or lead.customer_type
                        lead.enquiry_type = row.get('Enquiry Type') or lead.enquiry_type
                        lead.enquiry_source = row.get('Enquiry Source') or lead.enquiry_source
                        lead.sub_enquiry_source = row.get('Sub Enquiry Source') or lead.sub_enquiry_source
                        lead.lead_date = pd.to_datetime(row.get('Lead Date')).date() if not pd.isna(row.get('Lead Date')) else lead.lead_date
                        lead.call_date = pd.to_datetime(row.get('Call Date')).date() if not pd.isna(row.get('Call Date')) else lead.call_date
                        lead.call_type = row.get('Call Type') or lead.call_type
                        lead.calling_status = row.get('Calling Status') or lead.calling_status
                        lead.interested_status = row.get('Interested Status') or lead.interested_status
                        lead.sub_calling_status = row.get('Sub Calling Status') or lead.sub_calling_status
                        lead.sub_sub_calling_status = row.get('Sub Sub Calling Status') or lead.sub_sub_calling_status
                        lead.select_bus = row.get('Select Business') or lead.select_bus
                        lead.buyer_type = row.get('Buyer Type') or lead.buyer_type
                        lead.lead_status = row.get('Lead Status') or lead.lead_status
                        lead.construction_level = row.get('Construction Level') or lead.construction_level
                        lead.name = row.get('Name') or lead.name
                        lead.alternative_number = row.get('Alternative Number') or lead.alternative_number
                        lead.email_id = row.get('Email ID') or lead.email_id
                        lead.address = row.get('Address') or lead.address
                        lead.landmark = row.get('Landmark') or lead.landmark
                        lead.brand = row.get('Brand') or lead.brand
                        lead.product = row.get('Product') or lead.product
                        lead.sub_product = row.get('Sub Product') or lead.sub_product
                        lead.state = row.get('State') or lead.state
                        lead.district = row.get('District') or lead.district
                        lead.zone = row.get('Zone') or lead.zone
                        lead.pin_code = row.get('Pincode') or lead.pin_code
                        lead.agent_name = row.get('Agent Name') or lead.agent_name
                        lead.order_qty = int(row.get('Order Qty')) if not pd.isna(row.get('Order Qty')) else lead.order_qty
                        lead.order_description = row.get('Order Description') or lead.order_description
                        lead.order_value = float(row.get('Order Value')) if not pd.isna(row.get('Order Value')) else lead.order_value
                        lead.customer_type_select = row.get('Customer Type Select') or lead.customer_type_select
                        lead.registration_status = row.get('Registration Status') or lead.registration_status
                        lead.remark = row.get('Remark') or lead.remark
                        lead.secure_url = row.get('Secure URL') or lead.secure_url
                        lead.seller_email_id = row.get('Seller Email ID') or lead.seller_email_id
                        lead.seller_phone_no = row.get('Seller Phone No') or lead.seller_phone_no
                        lead.seller_email_id_L2 = row.get('Seller Email ID L2') or lead.seller_email_id_L2
                        lead.seller_phone_no_L2 = row.get('Seller Phone No L2') or lead.seller_phone_no_L2
                        lead.lead_closer_status = row.get('Lead Closer Status') or lead.lead_closer_status
                        lead.lead_closer_status_new = row.get('Lead Closer Status New') or lead.lead_closer_status_new
                        lead.lead_close_date = pd.to_datetime(row.get('Lead Close Date')).date() if not pd.isna(row.get('Lead Close Date')) else lead.lead_close_date
                        lead.final_lead_close_date = pd.to_datetime(row.get('Final Lead Close Date')).date() if not pd.isna(row.get('Final Lead Close Date')) else lead.final_lead_close_date

                        if request.user.is_authenticated:
                            lead.updated_by = request.user

                        lead.save()
                        update_count += 1

                    # --- Create new lead if Unique Id missing or not found ---
                    else:
                        lead = LeadTable(
                            calling_number=calling_number,
                            customer_name=customer_name,
                            customer_type=row.get('Customer Type') or None,
                            enquiry_type=row.get('Enquiry Type') or None,
                            enquiry_source=row.get('Enquiry Source') or None,
                            sub_enquiry_source=row.get('Sub Enquiry Source') or None,
                            lead_date=pd.to_datetime(row.get('Lead Date')).date() if not pd.isna(row.get('Lead Date')) else None,
                            call_date=pd.to_datetime(row.get('Call Date')).date() if not pd.isna(row.get('Call Date')) else None,
                            call_type=row.get('Call Type') or None,
                            calling_status=row.get('Calling Status') or None,
                            interested_status=row.get('Interested Status') or None,
                            sub_calling_status=row.get('Sub Calling Status') or None,
                            sub_sub_calling_status=row.get('Sub Sub Calling Status') or None,
                            select_bus=row.get('Select Business') or None,
                            buyer_type=row.get('Buyer Type') or None,
                            lead_status=row.get('Lead Status') or None,
                            construction_level=row.get('Construction Level') or None,
                            name=row.get('Name') or None,
                            alternative_number=row.get('Alternative Number') or None,
                            email_id=row.get('Email ID') or None,
                            address=row.get('Address') or None,
                            landmark=row.get('Landmark') or None,
                            brand=row.get('Brand') or None,
                            product=row.get('Product') or None,
                            sub_product=row.get('Sub Product') or None,
                            state=row.get('State') or None,
                            district=row.get('District') or None,
                            zone=row.get('Zone') or None,
                            pin_code=row.get('Pincode') or None,
                            agent_name=row.get('Agent Name') or None,
                            order_qty=int(row.get('Order Qty')) if not pd.isna(row.get('Order Qty')) else None,
                            order_description=row.get('Order Description') or None,
                            order_value=float(row.get('Order Value')) if not pd.isna(row.get('Order Value')) else None,
                            customer_type_select=row.get('Customer Type Select') or None,
                            registration_status=row.get('Registration Status') or None,
                            remark=row.get('Remark') or None,
                            secure_url=row.get('Secure URL') or None,
                            seller_email_id=row.get('Seller Email ID') or None,
                            seller_phone_no=row.get('Seller Phone No') or None,
                            seller_email_id_L2=row.get('Seller Email ID L2') or None,
                            seller_phone_no_L2=row.get('Seller Phone No L2') or None,
                            lead_closer_status=row.get('Lead Closer Status') or None,
                            lead_closer_status_new=row.get('Lead Closer Status New') or None,
                            lead_close_date=pd.to_datetime(row.get('Lead Close Date')).date() if not pd.isna(row.get('Lead Close Date')) else None,
                            final_lead_close_date=pd.to_datetime(row.get('Final Lead Close Date')).date() if not pd.isna(row.get('Final Lead Close Date')) else None,
                            lead_upload_type="Bulk"
                        )

                        if request.user.is_authenticated:
                            lead.created_by = request.user
                            lead.updated_by = request.user

                        lead.save()
                        success_count += 1

                except Exception as e:
                    failed_entries.append({
                        'Customer Name': customer_name or 'Unknown',
                        'reason': str(e)
                    })

            return JsonResponse({
                'status': 'completed',
                'created': success_count,
                'updated': update_count,
                'failed': len(failed_entries),
                'failed_entries': failed_entries
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'fail', 'message': 'Invalid request'}, status=400)



def download_excel_template_update(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads Update Template"

    headers = [
        'Unique Id','Customer Name', 'Customer Type', 'Calling Number', 'Enquiry Type', 'Enquiry Source',
        'Sub Enquiry Source', 'Lead Date', 'Call Date', 'Call Type', 'Calling Status', 'Interested Status',
        'Sub Calling Status', 'Sub Sub Calling Status', 'Select Business', 'Buyer Type', 'Lead Status',
        'Construction Level', 'Name', 'Alternative Number', 'Email ID', 'Address', 'Landmark', 'Brand',
        'Product', 'Sub Product', 'State', 'District', 'Zone', 'Pincode', 'Agent Name', 'Order Qty',
        'Order Description', 'Order Value', 'Customer Type Select', 'Registration Status', 'Remark',
        'Secure URL', 'Seller Email ID', 'Seller Phone No', 'Seller Email ID L2', 'Seller Phone No L2',
        'Lead Closer Status', 'Lead Closer Status New', 'Lead Close Date', 'Final Lead Close Date',
        'Created By', 'Updated By', 'Created At', 'Updated At'
    ]
    ws.append(headers)

    # Optional sample row
    sample_row = [
        'BN1','Sachin Kumar', 'Self', '9876543210', 'Product Inquiry', 'Website', 'Sub Source',
        '2025-07-14', '2025-07-14', 'Inbound', 'Connected', 'Interested', 'Follow-up', 'Final',
        'Bus A', 'Retail', 'Open', 'Level 1', 'Sachin', '9876543210', 'sachin@example.com',
        'Some Address', 'Near Park', 'BrandX', 'ProductY', 'SubProductZ', 'State1', 'District1',
        'Zone1', '123456', 'AgentX', 10, 'Sample Order', 1000.00, 'Type1', 'Registered', 'No remark',
        'http://secureurl.com', 'seller@example.com', '9876543210', 'seller2@example.com', '9876501234',
        'Closed', 'Closed New', '2025-07-20', '2025-07-21', 'admin', 'admin', '2025-07-14 10:00', '2025-07-14 10:30'
    ]
    ws.append(sample_row)

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="lead_update_template.xlsx"'
    return response
